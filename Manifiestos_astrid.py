# -*- coding: utf-8 -*-
"""
Created on Thu Feb 19 20:28:41 2026

@author: User
"""

# streamlit_app.py
# -*- coding: utf-8 -*-

import io
import pandas as pd
from datetime import datetime
from zoneinfo import ZoneInfo
import dropbox
import streamlit as st
import numpy as np
import json
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openai import OpenAI
import random
from faker import Faker

st.set_page_config(page_title="Manifiestos Astrid", layout="wide")

# -----------------------------
# LOGIN (antes de cargar la app)
# -----------------------------
def require_login():
    if "authed" not in st.session_state:
        st.session_state.authed = False

    if st.session_state.authed:
        return

    st.title("🔒 Login")

    pw = st.text_input("Contraseña", type="password")

    if st.button("Entrar"):
        if pw == st.secrets["auth"]["password"]:
            st.session_state.authed = True
            st.success("Acceso concedido ✅")
            st.rerun()
        else:
            st.error("Contraseña incorrecta ❌")

    st.stop()  # corta la app aquí si no está autenticado


require_login()





# Ruta EXACTA en Dropbox (siempre empieza con /)
DBX_FILE_PATH = "/Manifiestos/Manifiestos_astrid.xlsx"


# -----------------------------
# Dropbox client
# -----------------------------
def get_dbx() -> dropbox.Dropbox:
    cfg_dbx = st.secrets["dropbox"]
    return dropbox.Dropbox(
        app_key=cfg_dbx["app_key"],
        app_secret=cfg_dbx["app_secret"],
        oauth2_refresh_token=cfg_dbx["refresh_token"],
    )


# -----------------------------
# Helpers
# -----------------------------
def _clean_str_series(s: pd.Series) -> pd.Series:
    return (
        s.astype("string")
        .str.strip()
        .str.replace(r"\.0$", "", regex=True)
    )


def _norm_casillero(x) -> str:
    if pd.isna(x):
        return ""
    s = str(x).strip().upper()
    digits = "".join(ch for ch in s if ch.isdigit())
    return f"CA{digits}" if digits else s


def _clean_optional_text(x) -> str:
    if x is None:
        return ""
    return str(x).strip()


def _build_destinatario_real(nombre: str, direccion: str, telefono: str, ciudad: str, estado: str) -> dict:
    return {
        "NOMBRE DESTINO": _clean_optional_text(nombre),
        "DESTINO DIRECCION": _clean_optional_text(direccion),
        "DESTINO TELEFONO": _clean_optional_text(telefono),
        "DESTINO CIUDAD": _clean_optional_text(ciudad),
        "DESTINO ESTADO": _clean_optional_text(estado),
    }


def _build_destinatario_fake(fake: Faker) -> dict:
    ciudades_departamentos = [
        ("Medellín", "Antioquia"),
        ("Envigado", "Antioquia"),
        ("Rionegro", "Antioquia"),
        ("Cali", "Valle del Cauca"),
        ("Palmira", "Valle del Cauca"),
        ("Barranquilla", "Atlántico"),
        ("Bucaramanga", "Santander"),
        ("Floridablanca", "Santander"),
        ("Cartagena", "Bolívar"),
        ("Pereira", "Risaralda"),
        ("Manizales", "Caldas"),
        ("Armenia", "Quindío"),
        ("Ibagué", "Tolima"),
        ("Santa Marta", "Magdalena"),
        ("Montería", "Córdoba"),
        ("Neiva", "Huila"),
        ("Pasto", "Nariño"),
        ("Cúcuta", "Norte de Santander"),
        ("Soacha", "Cundinamarca"),
        ("Villavicencio", "Meta"),
    ]
    tipos_via = ["Calle", "Carrera", "Diagonal", "Transversal", "Avenida"]

    ciudad, departamento = random.choice(ciudades_departamentos)
    tipo_via = random.choice(tipos_via)
    numero_1 = random.randint(1, 120)
    numero_2 = random.randint(1, 99)
    numero_3 = random.randint(1, 99)
    telefono = "3" + "".join(random.choices("0123456789", k=9))
    direccion = f"{tipo_via} {numero_1} # {numero_2}-{numero_3}"

    return {
        "NOMBRE DESTINO": fake.name(),
        "DESTINO DIRECCION": direccion,
        "DESTINO TELEFONO": telefono,
        "DESTINO CIUDAD": ciudad,
        "DESTINO ESTADO": departamento,
    }


def apply_contingencia_destinatario(df: pd.DataFrame, casillero_objetivo: str, destinatario_real: dict) -> tuple[pd.DataFrame, int]:
    df = df.copy()
    df["_ES_CONTINGENCIA"] = False
    df["CONTINGENCIA_REAL"] = False

    casillero_norm = _norm_casillero(casillero_objetivo)
    mask_casillero = df["CASILLERO"].apply(_norm_casillero) == casillero_norm
    idx_objetivo = df.index[mask_casillero].tolist()

    if not idx_objetivo:
        return df, 0

    df.loc[idx_objetivo, "_ES_CONTINGENCIA"] = True

    idx_primero = idx_objetivo[0]
    df.at[idx_primero, "CONTINGENCIA_REAL"] = True
    for col, value in destinatario_real.items():
        df.at[idx_primero, col] = value

    fake = Faker("es_CO")
    for idx in idx_objetivo[1:]:
        destinatario_fake = _build_destinatario_fake(fake)
        for col, value in destinatario_fake.items():
            df.at[idx, col] = value

    return df, len(idx_objetivo)


def _normalize_excel_key(value) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip().removesuffix(".0")


def _paint_rows_red_by_guia(ws, guia_col_name: str, guias_rojas: set[str]) -> None:
    if not guias_rojas or ws.max_row < 2:
        return

    header_map = {cell.value: cell.column for cell in ws[1] if cell.value is not None}
    guia_col_idx = header_map.get(guia_col_name)
    if guia_col_idx is None:
        return

    red_fill = PatternFill(fill_type="solid", fgColor="FF4D4F")
    white_font = Font(color="FFFFFF", bold=True)

    for row_idx in range(2, ws.max_row + 1):
        guia_value = _normalize_excel_key(ws.cell(row=row_idx, column=guia_col_idx).value)
        if guia_value not in guias_rojas:
            continue
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = red_fill
            cell.font = white_font


def _hide_excel_columns(ws, column_names: list[str]) -> None:
    if not column_names or ws.max_row < 1:
        return

    header_map = {cell.value: cell.column for cell in ws[1] if cell.value is not None}
    for column_name in column_names:
        col_idx = header_map.get(column_name)
        if col_idx is not None:
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True


def dfs_to_excel_bytes(
    sheets: dict,
    highlight_rows: dict | None = None,
    hidden_columns: dict | None = None,
) -> bytes:
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for name, df in sheets.items():
            safe_name = str(name)[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)

            ws = writer.sheets[safe_name]
            config_highlight = (highlight_rows or {}).get(safe_name)
            if config_highlight:
                _paint_rows_red_by_guia(ws, config_highlight["guia_col"], config_highlight["guias"])

            cols_to_hide = (hidden_columns or {}).get(safe_name, [])
            if cols_to_hide:
                _hide_excel_columns(ws, cols_to_hide)

    out.seek(0)
    return out.getvalue()


# -----------------------------
# Formato INFO MANIFIESTO (Envios Encargomio / cajas)
# -----------------------------
INFO_MANIFIESTO_COLS = [
    "MASTER", "FECHA GUIA", "GUIA / consignment", "COMPAÑÍA REMITENTE",
    "REMITENTE DIRECCION", "REMITENTE CIUDAD", "REMITENTE ESTADO",
    "NOMBRE DESTINO", "DESTINO DIRECCION", "DESTINO CIUDAD", "CONTENIDO",
    "PESO LIBRAS", "PESO KILOS", "VALOR DECLARADO", "PIEZAS",
    "DESTINO ESTADO", "POSICION ARANCELARIA", "MANIFIESTO",
]


def build_info_manifiesto_df(df: pd.DataFrame) -> pd.DataFrame:
    """Reordena/renombra un manifiesto al formato INFO MANIFIESTO (cajas).
    MASTER queda vacía y 'guia' pasa a 'GUIA / consignment'."""
    d = df.copy()
    if "GUIA / consignment" not in d.columns:
        if "guia" in d.columns:
            d = d.rename(columns={"guia": "GUIA / consignment"})
        elif "GUIA" in d.columns:
            d = d.rename(columns={"GUIA": "GUIA / consignment"})
    d["MASTER"] = ""
    for c in INFO_MANIFIESTO_COLS:
        if c not in d.columns:
            d[c] = pd.NA
    return d[INFO_MANIFIESTO_COLS].copy()


def info_manifiesto_excel_bytes(df_info: pd.DataFrame) -> bytes:
    """Escribe el formato INFO MANIFIESTO y agrega el pie de TOTALES
    (W.Lb, W.Kg, Value, Pieces) igual que la hoja 'CAJAS ENCARGOMIO'."""
    def _sum(col):
        return round(float(pd.to_numeric(df_info[col], errors="coerce").sum()), 2)

    tot_lb = _sum("PESO LIBRAS")
    tot_kg = _sum("PESO KILOS")
    tot_val = _sum("VALOR DECLARADO")
    tot_pz = int(pd.to_numeric(df_info["PIEZAS"], errors="coerce").sum())
    idx = {name: i + 1 for i, name in enumerate(INFO_MANIFIESTO_COLS)}

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df_info.to_excel(writer, sheet_name="INFO MANIFIESTO", index=False)
        ws = writer.sheets["INFO MANIFIESTO"]

        r = ws.max_row + 2          # fila de valores (deja una en blanco tras los datos)
        lab = r + 1                 # fila de leyenda
        ws.cell(r, idx["PESO LIBRAS"]).value = tot_lb
        ws.cell(r, idx["PESO KILOS"]).value = tot_kg
        ws.cell(r, idx["VALOR DECLARADO"]).value = tot_val
        ws.cell(r, idx["PIEZAS"]).value = tot_pz

        ws.cell(lab, idx["MASTER"]).value = "TOTAL:"
        ws.cell(lab, idx["DESTINO CIUDAD"]).value = "Volume Vlb:0"
        ws.cell(lab, idx["PESO LIBRAS"]).value = "W.Lb"
        ws.cell(lab, idx["PESO KILOS"]).value = "W.Kg:"
        ws.cell(lab, idx["VALOR DECLARADO"]).value = "Value:$"
        ws.cell(lab, idx["PIEZAS"]).value = "Pieces:"

        for cell in ws[r]:
            cell.font = Font(bold=True)
        for cell in ws[lab]:
            cell.font = Font(bold=True)
    out.seek(0)
    return out.getvalue()


# -----------------------------
# Clasificación CONTENIDO por IA (OpenAI) — usada por Envios Encargomio
# (Manifiestos Luma define su propia copia local; esta es a nivel de módulo)
# -----------------------------
CATEGORIAS_IA = [
    "Tenis", "Calzado", "Celular", "Computador", "Componente de computador",
    "Ropa", "Perfumes", "Cosmeticos", "Accesorios", "Reloj/Joyeria",
    "Hogar", "Electrodomestico", "Juguetes", "Herramientas",
    "Suplementos", "Medicamentos", "Alimentos",
    "Libros/Papeleria", "Miscelaneo",
]

CLASIFICAR_ENVIO_SCHEMA_IA = {
    "type": "object",
    "properties": {
        "categoria": {"type": "string", "enum": CATEGORIAS_IA},
        "confianza": {"type": "integer", "minimum": 0, "maximum": 100},
        "contenido": {
            "type": "string",
            "description": "Texto corto para CONTENIDO (sin marcas). Ej: 'Calzado (tenis)', 'Celular', 'Ropa'.",
        },
    },
    "required": ["categoria", "confianza", "contenido"],
    "additionalProperties": False,
}


def get_openai_client_mod() -> OpenAI:
    return OpenAI(api_key=st.secrets["openai"]["api_key"])


def gpt_clasificar_envio_mod(client: OpenAI, producto_dominante: str, productos_lista: str) -> dict:
    prompt = f"""
Eres un clasificador de envíos para manifiestos.
Objetivo: llenar la columna CONTENIDO con una categoria GENERAL, sin marcas.

Reglas:
- Prioriza PRODUCTO_DOMINANTE (es el más pesado del envío).
- Usa PRODUCTOS_LISTA solo como contexto.
- NO menciones marcas (Nike, Apple, etc). Solo categoria general.
- 'contenido' debe ser corto (1-4 palabras). Puedes aclarar entre paréntesis sin marcas.
- Si es muy variado, usa Miscelaneo.

PRODUCTO_DOMINANTE: {producto_dominante}
PRODUCTOS_LISTA: {productos_lista}
""".strip()

    resp = client.responses.create(
        model="gpt-4o-mini",
        input=prompt,
        text={
            "format": {
                "type": "json_schema",
                "name": "clasificar_envio",
                "schema": CLASIFICAR_ENVIO_SCHEMA_IA,
                "strict": True,
            }
        },
    )
    return json.loads(resp.output_text)


# Categorías "no útiles" de Encargomio: se tratan como vacías y se reclasifican
# con IA (la DIAN no admite "Otro"). Se comparan en minúsculas por substring.
CATEGORIAS_GENERICAS = ("otro", "category for import", "sin categoria", "sin categoría")


def _es_categoria_generica(v) -> bool:
    if v is None or pd.isna(v):
        return True
    s = str(v).strip().lower()
    if s in ("", "nan"):
        return True
    return any(g in s for g in CATEGORIAS_GENERICAS)


def enriquecer_contenido_ia(df: pd.DataFrame, df_p: pd.DataFrame, guia_col: str = "guia", solo_guias=None) -> tuple[pd.DataFrame, int, list]:
    """Reclasifica CONTENIDO vacío o genérico ('Otro', etc.) en df (llave guia_col)
    con IA usando el producto dominante del casillero (df_p). Las que vienen 'Otro'/
    vacío y NO tienen producto (no se pueden clasificar) se dejan EN BLANCO y se
    marcan para categoría manual (la DIAN no admite 'Otro'). Si solo_guias se pasa,
    ese blanqueo/marcado se limita a esas guías (el lote nuevo), para no alterar el
    histórico viejo. Devuelve (df, n_clasificadas, guias_sin_categoria)."""
    df = df.copy()
    COL_ENVIO, COL_PROD, COL_PESO = "Envío", "Nombre producto", "Peso"
    if any(c not in df_p.columns for c in (COL_ENVIO, COL_PROD, COL_PESO)):
        return df, 0

    dp = df_p[[COL_ENVIO, COL_PROD, COL_PESO]].copy()
    dp[COL_ENVIO] = _clean_str_series(dp[COL_ENVIO])
    dp[COL_PROD] = dp[COL_PROD].astype("string").fillna("").str.strip()
    dp[COL_PESO] = pd.to_numeric(dp[COL_PESO], errors="coerce").fillna(0)
    dp = dp[dp[COL_PROD] != ""]
    if dp.empty:
        return df, 0

    dp = dp.groupby([COL_ENVIO, COL_PROD], as_index=False)[COL_PESO].sum()
    idx_dom = dp.groupby(COL_ENVIO)[COL_PESO].idxmax()
    dom = dp.loc[idx_dom, [COL_ENVIO, COL_PROD]].rename(columns={COL_PROD: "_DOM"})
    lst = (
        dp.sort_values([COL_ENVIO, COL_PESO], ascending=[True, False])
        .groupby(COL_ENVIO, as_index=False)[COL_PROD]
        .apply(lambda s: " | ".join(s.head(50).tolist()))
        .rename(columns={COL_PROD: "_LISTA"})
    )
    agg = dom.merge(lst, on=COL_ENVIO, how="left")

    if "CONTENIDO" not in df.columns:
        df["CONTENIDO"] = pd.NA
    df["CONTENIDO"] = df["CONTENIDO"].astype("string")

    key = _clean_str_series(df[guia_col])
    m = pd.DataFrame({"_k": key.values}).merge(agg, how="left", left_on="_k", right_on=COL_ENVIO)
    df["_DOM"] = m["_DOM"].values
    df["_LISTA"] = m["_LISTA"].values

    # "Necesita categoría" = vacío o genérico ('Otro', 'Category for import', ...)
    generica = df["CONTENIDO"].map(_es_categoria_generica)
    tiene_dom = df["_DOM"].notna() & (df["_DOM"].astype(str).str.strip() != "")

    # Reclasificar SOLO las que tienen producto en el casillero (para no borrar
    # datos de guías viejas que no vienen en este reporte de productos).
    mask = generica & tiene_dom
    n = int(mask.sum())
    if n:
        client = get_openai_client_mod()
        cache, conts = {}, []
        for d_, l_ in zip(
            df.loc[mask, "_DOM"].astype(str).tolist(),
            df.loc[mask, "_LISTA"].fillna("").astype(str).tolist(),
        ):
            l_ = l_[:3000]
            k = (d_, l_)
            if k not in cache:
                cache[k] = gpt_clasificar_envio_mod(client, d_, l_)
            conts.append(cache[k]["contenido"])
        df.loc[mask, "CONTENIDO"] = conts

    # Genéricas/vacías que NO se pudieron clasificar (sin producto en el casillero):
    # se dejan en BLANCO (nunca 'Otro') y se marcan para categoría manual (DIAN).
    candidatos = generica & ~tiene_dom
    if solo_guias is not None:
        candidatos = candidatos & _clean_str_series(df[guia_col]).isin(set(solo_guias))
    df.loc[candidatos, "CONTENIDO"] = pd.NA
    sin_categoria = df.loc[candidatos, guia_col].dropna().astype(str).tolist()

    return df.drop(columns=["_DOM", "_LISTA"]), n, sin_categoria


def procesar_envios_encargomio(df_a: pd.DataFrame, df_b: pd.DataFrame, hist: pd.DataFrame, df_p=None):
    """Cruza pistoleo (A) × Envíos Encargomio (B), arma columnas, concatena con el
    histórico propio (dedup por guia, histórico manda), numera manifiestos desde
    2000001 (+1 por corrida) y clasifica CONTENIDO vacío con IA si se pasa df_p.
    Devuelve (df_concat, nuevo_manifiesto, n_ia). n_ia = -1 si la IA falló."""
    MAN_BASE = 2000001

    df_a = df_a.copy()
    df_b = df_b.copy()
    df_a["Envio"] = _clean_str_series(df_a["Envio"])
    df_a = df_a.drop_duplicates(subset=["Envio"], keep="first")
    df_b["NUMERO ENVIO"] = _clean_str_series(df_b["NUMERO ENVIO"])

    alias_cols = {
        "DIRECCIÓN DESTINO": "DESTINO DIRECCION",
        "DIRECCIÃ“N DESTINO": "DESTINO DIRECCION",
        "TELÉFONO": "DESTINO TELEFONO",
        "TELÃ‰FONO": "DESTINO TELEFONO",
    }
    for src, dst in alias_cols.items():
        if src in df_b.columns:
            df_b = df_b.rename(columns={src: dst})
    df_b = df_b.rename(columns={
        "CLIENTE DESTINO": "NOMBRE DESTINO",
        "CIUDAD DESTINO": "DESTINO CIUDAD",
        "DEPARTAMENTO DESTINO": "DESTINO ESTADO",
    })

    df_b["COMPAÑÍA REMITENTE"] = "Largo Easy Corp"
    df_b["REMITENTE DIRECCION"] = "11860 SW 144th Ct Ste 2"
    df_b["REMITENTE TELEFONO"] = "3053996614"
    df_b["REMITENTE CIUDAD"] = "Miami"
    df_b["REMITENTE ESTADO"] = "FL"

    df_a = df_a.rename(columns={"Envio": "guia"})
    df_b = df_b.rename(columns={"NUMERO ENVIO": "guia"})
    df_a["guia"] = _clean_str_series(df_a["guia"])
    df_b["guia"] = _clean_str_series(df_b["guia"])
    if "CATEGORÍAS PRODUCTOS" in df_b.columns:
        df_b = df_b.rename(columns={"CATEGORÍAS PRODUCTOS": "CONTENIDO"})

    cols_b = [
        "guia", "CASILLERO", "VALOR", "COMPAÑÍA REMITENTE", "REMITENTE DIRECCION",
        "REMITENTE CIUDAD", "REMITENTE ESTADO", "NOMBRE DESTINO", "DESTINO DIRECCION",
        "DESTINO CIUDAD", "DESTINO ESTADO", "CONTENIDO", "PESO",
    ]
    cols_b = [c for c in cols_b if c in df_b.columns]
    df_b_sel = df_b[cols_b].rename(columns={"PESO": "PESO LIBRAS"})

    df_final = df_a.merge(df_b_sel, how="left", on="guia")
    df_final["PIEZAS"] = 1
    df_final["VALOR DECLARADO"] = np.random.randint(91, 100, size=len(df_final))
    df_final["POSICION ARANCELARIA"] = "980720"
    df_final["PESO LIBRAS"] = pd.to_numeric(df_final["PESO LIBRAS"], errors="coerce")
    df_final["PESO KILOS"] = df_final["PESO LIBRAS"] / 2.20462
    df_final["FECHA GUIA"] = pd.to_datetime(datetime.now(ZoneInfo("America/New_York")).date())
    if "CONTENIDO" not in df_final.columns:
        df_final["CONTENIDO"] = pd.NA

    # Guías pistoleadas sin cruce en B / sin peso válido (informativo, no bloquea)
    _peso_new = df_final["PESO LIBRAS"]
    guias_sin_peso = (
        df_final.loc[_peso_new.isna() | (_peso_new <= 0), "guia"]
        .dropna()
        .astype(str)
        .tolist()
    )

    hist = hist.copy() if hist is not None else pd.DataFrame()
    if "guia" not in hist.columns and "GUIA" in hist.columns:
        hist = hist.rename(columns={"GUIA": "guia"})
    if not hist.empty and "guia" in hist.columns:
        hist["guia"] = _clean_str_series(hist["guia"])

    guias_hist_set = set(hist["guia"].dropna()) if ("guia" in hist.columns and not hist.empty) else set()
    guias_nuevas = set(df_final["guia"].dropna()) - guias_hist_set

    df_concat = pd.concat([hist, df_final], ignore_index=True)
    df_concat = df_concat.drop_duplicates(subset=["guia"], keep="first").reset_index(drop=True)

    if "MANIFIESTO" not in df_concat.columns:
        df_concat["MANIFIESTO"] = pd.NA
    man_num = pd.to_numeric(df_concat["MANIFIESTO"], errors="coerce")
    mx = man_num[man_num >= MAN_BASE].max()
    nuevo = int(mx) + 1 if pd.notna(mx) else MAN_BASE
    df_concat.loc[man_num.isna(), "MANIFIESTO"] = nuevo
    df_concat["MANIFIESTO"] = pd.to_numeric(df_concat["MANIFIESTO"], errors="coerce").astype("Int64")

    n_ia = 0
    sin_categoria = []
    if df_p is not None:
        try:
            df_concat, n_ia, sin_categoria = enriquecer_contenido_ia(
                df_concat, df_p, guia_col="guia", solo_guias=guias_nuevas
            )
        except Exception:
            n_ia = -1

    return df_concat, int(nuevo), n_ia, guias_sin_peso, sin_categoria


# -----------------------------
# UI: Uploaders
# -----------------------------
st.title("Manifiestos Astrid")

modo = st.radio(
    "Selecciona el proceso",
    ["Manifiestos Luma", "Celulares Fénix", "Envios Encargomio"],
    horizontal=True,
)

if modo == "Manifiestos Luma":
    col1, col2, col3 = st.columns(3)
    with col1:
        up_a = st.file_uploader("Sube Pistoleo de bodega (Envios pistoleo)", type=["xlsx", "xls"])
    with col2:
        up_b = st.file_uploader("Sube Envíos Encargomio (Envios_Encargomio)", type=["xlsx", "xls"])
    with col3:
        up_p = st.file_uploader("Sube Productos por casillero", type=["xlsx", "xls"])

    st.markdown("### Contingencia")
    contingencia_activa = st.toggle(
        "Activar contingencia para un casillero puntual",
        value=False,
        help="Deja el primer registro nuevo del casillero con destinatario real y simula los siguientes.",
    )

    cont_casillero = ""
    cont_nombre = ""
    cont_direccion = ""
    cont_telefono = ""
    cont_ciudad = ""
    cont_estado = ""
    contingencia_lista = True

    if contingencia_activa:
        st.info("La contingencia solo afecta guías nuevas del casillero indicado en esta corrida.")

        cont_col1, cont_col2, cont_col3 = st.columns(3)
        with cont_col1:
            cont_casillero = st.text_input("Casillero de contingencia")
            cont_nombre = st.text_input("Nombre destinatario real")
        with cont_col2:
            cont_direccion = st.text_input("Dirección destinatario real")
            cont_telefono = st.text_input("Teléfono destinatario real")
        with cont_col3:
            cont_ciudad = st.text_input("Ciudad destinatario real")
            cont_estado = st.text_input("Estado destinatario real")

        campos_contingencia = [
            cont_casillero,
            cont_nombre,
            cont_direccion,
            cont_telefono,
            cont_ciudad,
            cont_estado,
        ]
        contingencia_lista = all(_clean_optional_text(v) for v in campos_contingencia)

        if not contingencia_lista:
            st.warning("Completa casillero y todos los datos reales del destinatario para usar la contingencia.")

    run = st.button(
        "Procesar y actualizar histórico en Dropbox",
        type="primary",
        disabled=not (up_a and up_b and up_p and contingencia_lista)
    )
    if run:
        # -----------------------------
        # 1) Leer histórico de Dropbox
        # -----------------------------
        dbx = get_dbx()
        def load_historico_or_empty(dbx: dropbox.Dropbox, path: str) -> pd.DataFrame:
            # Columnas mínimas esperadas del histórico
            base_cols = [
                "FECHA GUIA", "guia", "COMPAÑÍA REMITENTE", "REMITENTE DIRECCION", "REMITENTE TELEFONO",
                "REMITENTE CIUDAD", "REMITENTE ESTADO", "NOMBRE DESTINO", "DESTINO DIRECCION",
                "DESTINO TELEFONO", "DESTINO CIUDAD", "CONTENIDO", "PESO LIBRAS", "PESO KILOS",
                "VALOR DECLARADO", "PIEZAS", "DESTINO ESTADO", "POSICION ARANCELARIA", "MANIFIESTO","INSTRUCCIONES",
                "CASILLERO", "VALOR"  # útil para lógica de manifiesto
            ]
            try:
                _, res = dbx.files_download(path)
                df = pd.read_excel(io.BytesIO(res.content), sheet_name=0)
                # asegurar columna guia si viene como GUIA u otra variante
                if "guia" not in df.columns and "GUIA" in df.columns:
                    df = df.rename(columns={"GUIA": "guia"})
                return df
            except dropbox.exceptions.ApiError as e:
                # si no existe, arrancar vacío
                if getattr(e, "error", None) and e.error.is_path() and e.error.get_path().is_not_found():
                    return pd.DataFrame(columns=base_cols)
                raise  # cualquier otro error sí lo mostramos
        df_historico = load_historico_or_empty(dbx, DBX_FILE_PATH)

        st.success(f"Histórico descargado: {df_historico.shape[0]} filas")

        # -----------------------------
        # 2) Leer archivos subidos (A y B)
        # -----------------------------
        df_a = pd.read_excel(up_a)
        df_b = pd.read_excel(up_b)

        st.info(f"A (Pistoleo): {df_a.shape} | B (Encargomio): {df_b.shape}")

        # -----------------------------
        # 3) Asegurar STR + limpieza .0 en llaves originales
        # -----------------------------
        df_a["Envio"] = _clean_str_series(df_a["Envio"])
        df_a = df_a.drop_duplicates(subset=["Envio"], keep="first").copy()
    
        df_b["NUMERO ENVIO"] = _clean_str_series(df_b["NUMERO ENVIO"])

        # -----------------------------
        # 4) Renombres remitente
        # -----------------------------
        rename_map = {
            "CLIENTE DESTINO": "NOMBRE DESTINO",
            "CIUDAD DESTINO": "DESTINO CIUDAD",
            "DEPARTAMENTO DESTINO": "DESTINO ESTADO",
        }
        alias_cols = {
            "DIRECCIÓN DESTINO": "DESTINO DIRECCION",
            "DIRECCIÃ“N DESTINO": "DESTINO DIRECCION",
            "TELÉFONO": "DESTINO TELEFONO",
            "TELÃ‰FONO": "DESTINO TELEFONO",
        }

        for src, dst in alias_cols.items():
            if src in df_b.columns:
                df_b = df_b.rename(columns={src: dst})

        faltantes = [c for c in rename_map.keys() if c not in df_b.columns]
        if faltantes:
            st.error(f"Faltan columnas en B: {faltantes}")
            st.stop()

        df_b = df_b.rename(columns=rename_map)

        faltantes_alias = [c for c in ["DESTINO DIRECCION", "DESTINO TELEFONO"] if c not in df_b.columns]
        if faltantes_alias:
            st.error(f"Faltan columnas en B: {faltantes_alias}")
            st.stop()

        # -----------------------------


    

        df_b["COMPAÑÍA REMITENTE"] =  "Largo Easy Corp"
        df_b["REMITENTE DIRECCION"] = "11860 SW 144th Ct Ste 2"
        df_b["REMITENTE TELEFONO"] = "3053996614"
        df_b["REMITENTE CIUDAD"] = "Miami"
        df_b["REMITENTE ESTADO"] = "FL"

    
    


        # -----------------------------
        # 6) Renombrar llaves a guia
        # -----------------------------
        df_a = df_a.rename(columns={"Envio": "guia"})
        df_b = df_b.rename(columns={"NUMERO ENVIO": "guia"})

        df_a["guia"] = _clean_str_series(df_a["guia"])
        df_b["guia"] = _clean_str_series(df_b["guia"])
        if "CATEGORÍAS PRODUCTOS" in df_b.columns:
            df_b = df_b.rename(columns={"CATEGORÍAS PRODUCTOS": "CONTENIDO"})

        # -----------------------------
        # 7) Selección columnas B + PESO -> PESO LIBRAS
        # -----------------------------
        cols_b = [
            "guia",
            "CASILLERO",
            "VALOR",
            "COMPAÑÍA REMITENTE",
            "REMITENTE DIRECCION",
            "REMITENTE TELEFONO",
            "REMITENTE CIUDAD",
            "REMITENTE ESTADO",
            "NOMBRE DESTINO",
            "DESTINO DIRECCION",
            "DESTINO TELEFONO",
            "DESTINO CIUDAD",
            "DESTINO ESTADO",
            "PESO",
        ]
        falt_b = [c for c in cols_b if c not in df_b.columns]
        if falt_b:
            st.error(f"Faltan columnas requeridas en B para el cruce: {falt_b}")
            st.stop()

        df_b_sel = df_b[cols_b].rename(columns={"PESO": "PESO LIBRAS"})

        # -----------------------------
        # 8) Merge
        # -----------------------------
        df_final = df_a.merge(df_b_sel, how="left", on="guia")
        df_final["_ES_CONTINGENCIA"] = False
        df_final["CONTINGENCIA_REAL"] = False

        if contingencia_activa:
            destinatario_real = _build_destinatario_real(
                cont_nombre,
                cont_direccion,
                cont_telefono,
                cont_ciudad,
                cont_estado,
            )

            # Aplicar contingencia solo sobre guías que no están en el histórico
            guias_en_historico = set(_clean_str_series(df_historico["guia"]).dropna())
            mask_nuevas = ~df_final["guia"].isin(guias_en_historico)

            df_nuevas = df_final[mask_nuevas].copy()
            df_viejas = df_final[~mask_nuevas].copy()

            df_nuevas, total_contingencia = apply_contingencia_destinatario(
                df_nuevas,
                cont_casillero,
                destinatario_real,
            )

            df_final = pd.concat([df_viejas, df_nuevas], ignore_index=True)

            if total_contingencia == 0:
                st.warning(
                    f"Contingencia activa, pero no se encontraron guías nuevas para el casillero {_norm_casillero(cont_casillero)}."
                )
            else:
                simuladas = max(total_contingencia - 1, 0)
                st.info(
                    f"Contingencia aplicada a {total_contingencia} guía(s) nuevas de {_norm_casillero(cont_casillero)}: 1 real y {simuladas} simulada(s)."
                )

        # -----------------------------
    # 8.1) Columnas nuevas
    # -----------------------------
        df_final["PIEZAS"] = 1
        df_final["VALOR DECLARADO"] = np.random.randint(91, 100, size=len(df_final))  # 91–99
        df_final["POSICION ARANCELARIA"] = "980720"
        df_final["POSICION ARANCELARIA"] = df_final["POSICION ARANCELARIA"].astype("string")

        # -----------------------------
        # 9) PESO KILOS + FECHA GUIA Miami
        # -----------------------------
        df_final["PESO LIBRAS"] = pd.to_numeric(df_final["PESO LIBRAS"], errors="coerce")
        df_final["PESO KILOS"] = df_final["PESO LIBRAS"] / 2.20462

        hoy_miami = datetime.now(ZoneInfo("America/New_York")).date()
        df_final["FECHA GUIA"] = pd.to_datetime(hoy_miami)


        # -----------------------------
        # 9.1) COSTO en filas nuevas (para validar antes de subir)
        # -----------------------------
        w_new = df_final["PESO LIBRAS"]
    
        df_final["COSTO"] = np.where(
            w_new.isna() | (w_new <= 0),
            0,
            np.where(
                w_new == 1,
                7.25,
                7.25 + (w_new - 1) * 2.6
            ) + np.where(w_new >= 20, 4.75, 0)
        ).round(2)
    
    
    
        # -----------------------------
    # -----------------------------
        # 9.2) VALIDACIÓN BLOQUEANTE
        # Si una guía nueva no encontró peso/costo, se detiene el proceso
        # EXCEPTO guías permitidas
        # -----------------------------
    
        GUIAS_EXCLUIDAS_ALERTA = {"85674", "8646","82760","83685","84147","84243","84554","84657","84670","84722","84457","84669","84663","84540","84827","84862","84824","84793","84806","84826","84818","84840","84950","84916","84874","84940","84956","84958","85029","84964","85024","85043","85048","85017","85046","85016","85008","85006","85036","85191","85194","85130","85169","85136","85152","85159","85161","85157","85173","85126","85125","85142","85140","85154","85155","85156","85128","85138","85139","85097","85187","85188","85186","85085","85084","85121","85163","85133","85141","85120","85093","85211","85210","85206","85204","85197","85201","85202","85203","85212","85198","85199","85205","85200","85208","85321","85322","85320","85317","85319","85318","85314","85312","85313","85316","85244","85333","85340","85302","85269","85304","85275","85280","85273","85266","85271","85267","85256","85251","85233","85341","85331","85332","85327","85342","85344","85240","85227","85292","85305","85264","85254","85265","85277","85282","85228","85260","85261","85259","85278","85241","85270","85229","85272","85345","85378","85368","85346","85360","85362","85377","85374","85347","85348","85375","85372","85400","85398","85349","85351","85370","85393","85406","85423","85427","85425","85424","85428","85426","85402","85421","85413","85422","85420","85382","85411","85404","85405","85401","85412","85449","85445","85446","85444","85441","85443","85450","85431","85373","85410","85379","85392","85350","85354","85353","85387","85367"}
    
        # normalizar guia como texto limpio
        df_final["guia"] = _clean_str_series(df_final["guia"])

        mask_error_nuevos = (
            df_final["PESO LIBRAS"].isna()
            | (df_final["PESO LIBRAS"] <= 0)
        )

        # guías ya registradas en histórico con FECHA GUIA > 4 días → excluir de alerta
        limite_alerta = pd.Timestamp(hoy_miami) - pd.Timedelta(days=4)
        guias_viejas_historico = set()
        if "FECHA GUIA" in df_historico.columns and not df_historico.empty:
            fechas_hist = pd.to_datetime(df_historico["FECHA GUIA"], errors="coerce")
            mask_viejas = fechas_hist < limite_alerta
            guias_viejas_historico = set(
                _clean_str_series(df_historico.loc[mask_viejas, "guia"]).dropna()
            )

        mask_excluidas = (
            df_final["guia"].isin(GUIAS_EXCLUIDAS_ALERTA)
            | df_final["guia"].isin(guias_viejas_historico)
        )
    
        df_errores_nuevos = df_final.loc[
            mask_error_nuevos & ~mask_excluidas,
            ["guia", "PESO LIBRAS", "COSTO"]
        ].copy()
    
        if not df_errores_nuevos.empty:
            st.error("❌ Proceso detenido. Hay guías nuevas pistoleadas que no fueron encontradas correctamente en Envíos Encargomio.")
    
            st.warning(
                "Revisa las siguientes guías pistoleadas. "
                "Tienen PESO LIBRAS vacío, menor o igual a 0, o COSTO en 0. "
                "Por eso el histórico no se cargará."
            )
    
            st.dataframe(df_errores_nuevos, use_container_width=True)
    
            guias_error = df_errores_nuevos["guia"].dropna().astype(str).tolist()
            st.markdown("**Guías con problema:**")
            st.write(", ".join(guias_error))
    
            st.stop()



        # -----------------------------
        # 10) Concat + dedup (histórico manda)
        # -----------------------------
        df_historico["guia"] = _clean_str_series(df_historico["guia"])
        df_final["guia"] = _clean_str_series(df_final["guia"])
        df_historico["_ES_CONTINGENCIA"] = False
        df_final["_ES_CONTINGENCIA"] = df_final["_ES_CONTINGENCIA"].fillna(False)
        if "CONTINGENCIA_REAL" not in df_historico.columns:
            df_historico["CONTINGENCIA_REAL"] = False
        df_historico["CONTINGENCIA_REAL"] = df_historico["CONTINGENCIA_REAL"].fillna(False)
        df_final["CONTINGENCIA_REAL"] = df_final["CONTINGENCIA_REAL"].fillna(False)

        df_concat = pd.concat([df_historico, df_final], ignore_index=True)
        df_concat = df_concat.drop_duplicates(subset=["guia"], keep="first").reset_index(drop=True)

        # -----------------------------
        # 11) Crear MANIFIESTO solo a vacíos con regla 11591 vs otros
        # -----------------------------
        if "MANIFIESTO" not in df_concat.columns:
            df_concat["MANIFIESTO"] = pd.NA

        if "_ES_CONTINGENCIA" not in df_concat.columns:
            df_concat["_ES_CONTINGENCIA"] = False

        df_concat["_ES_CONTINGENCIA"] = df_concat["_ES_CONTINGENCIA"].fillna(False).astype(bool)
        df_concat["CASILLERO_NORM"] = df_concat["CASILLERO"].apply(_norm_casillero)
        df_concat["MANIFIESTO_NUM"] = pd.to_numeric(df_concat["MANIFIESTO"], errors="coerce")

        mask_vacio = df_concat["MANIFIESTO_NUM"].isna()
        mask_contingencia = df_concat["_ES_CONTINGENCIA"]
        mask_11591 = df_concat["CASILLERO_NORM"] == "CA11591"
        man_num_str = df_concat["MANIFIESTO_NUM"].astype("Int64").astype("string")

        max_11591 = df_concat.loc[~mask_vacio & mask_11591 & ~mask_contingencia, "MANIFIESTO_NUM"].max()
        max_otros = df_concat.loc[~mask_vacio & ~mask_11591 & ~mask_contingencia, "MANIFIESTO_NUM"].max()
        max_contingencia = df_concat.loc[
            df_concat["MANIFIESTO_NUM"].notna() & man_num_str.str.startswith("4").fillna(False),
            "MANIFIESTO_NUM"
        ].max()

        if pd.isna(max_11591):
            max_11591 = 900000
        if pd.isna(max_otros):
            max_otros = 100000
        if pd.isna(max_contingencia):
            max_contingencia = 400000

        nuevo_man_11591 = int(max_11591) + 1
        nuevo_man_otros = int(max_otros) + 1
        nuevo_man_contingencia = int(max_contingencia) + 1 if mask_contingencia.any() else None

        if nuevo_man_contingencia is not None:
            df_concat.loc[mask_vacio & mask_contingencia, "MANIFIESTO"] = nuevo_man_contingencia
        df_concat.loc[mask_vacio & mask_11591 & ~mask_contingencia, "MANIFIESTO"] = nuevo_man_11591
        df_concat.loc[mask_vacio & ~mask_11591 & ~mask_contingencia, "MANIFIESTO"] = nuevo_man_otros

        df_concat["MANIFIESTO"] = pd.to_numeric(df_concat["MANIFIESTO"], errors="coerce").astype("Int64")
        df_concat = df_concat.drop(columns=["CASILLERO_NORM", "MANIFIESTO_NUM", "_ES_CONTINGENCIA"], errors="ignore")
        # -----------------------------
        # 11.5) Reordenar columnas para export
        # -----------------------------
        orden_cols = [
            "FECHA GUIA",
            "guia",
            "COMPAÑÍA REMITENTE",
            "REMITENTE DIRECCION",
            "REMITENTE TELEFONO",
            "REMITENTE CIUDAD",
            "REMITENTE ESTADO",
            "NOMBRE DESTINO",
            "DESTINO DIRECCION",
            "DESTINO TELEFONO",
            "DESTINO CIUDAD",
            "CONTENIDO",
            "PESO LIBRAS",
            "PESO KILOS",
            "VALOR DECLARADO",
            "PIEZAS",
            "DESTINO ESTADO",
            "POSICION ARANCELARIA",
            "MANIFIESTO",
            "INSTRUCCIONES",
            "CASILLERO",
            "VALOR"
        ]
    
        # solo deja las que existan (por si el histórico viejo trae extras)
        presentes = [c for c in orden_cols if c in df_concat.columns]
        extras = [c for c in df_concat.columns if c not in presentes]
        df_concat = df_concat[presentes + extras]
        # -----------------------------
        # X) COSTO (se guarda en Dropbox)
        # -----------------------------
        df_concat["PESO LIBRAS"] = pd.to_numeric(df_concat["PESO LIBRAS"], errors="coerce")
        w = df_concat["PESO LIBRAS"]
    
        df_concat["COSTO"] = np.where(
            w.isna() | (w <= 0),
            0,
            np.where(
                w == 1,
                7.25,
                7.25 + (w - 1) * 2.6
            ) + np.where(w >= 20, 4.75, 0)
        ).round(2)
    
        # -----------------------------
        # X) Resumen COSTO por manifiesto (para hoja 2)
        # -----------------------------
        # Asegurar numéricos
        for col in ["PESO LIBRAS", "PESO KILOS", "PIEZAS", "COSTO"]:
            if col in df_concat.columns:
                df_concat[col] = pd.to_numeric(df_concat[col], errors="coerce")
    
        df_costos_manifiesto = (
            df_concat
            .groupby("MANIFIESTO", dropna=False, as_index=False)[["PESO LIBRAS", "PESO KILOS", "PIEZAS", "COSTO"]]
            .sum(min_count=1)
        )
    
        # (opcional) ordenar
        df_costos_manifiesto = df_costos_manifiesto.sort_values("MANIFIESTO", na_position="last")
    
    
    
    
        # =========================
        # PRODUCTOS -> PRODUCTOS_RAW (solo si CONTENIDO está vacío)
        # =========================
    
        # 0) Cargar archivo productos (si ya lo tienes, omite esta línea)
    # =========================
    # PRODUCTOS (df_p) -> resumen por Envio usando Peso
    # y cruce SOLO para filas sin CONTENIDO
    # =========================

    # df_p ya leído (si no:)
        df_p = pd.read_excel(up_p)
    
        COL_ENVIO = "Envío"
        COL_PROD  = "Nombre producto"
        COL_PESO  = "Peso"
    
        df_prod = df_p[[COL_ENVIO, COL_PROD, COL_PESO]].copy()
    
        df_prod[COL_ENVIO] = _clean_str_series(df_prod[COL_ENVIO])
        df_prod[COL_PROD]  = df_prod[COL_PROD].astype("string").fillna("").str.strip()
        df_prod[COL_PESO]  = pd.to_numeric(df_prod[COL_PESO], errors="coerce").fillna(0)
    
        df_prod = df_prod[df_prod[COL_PROD] != ""].copy()
    
        # ✅ DEDUP real: colapsa repetidos (Envio, Producto) sumando Peso
        df_prod = (
            df_prod
            .groupby([COL_ENVIO, COL_PROD], as_index=False)[COL_PESO]
            .sum()
        )
    
        # Dominante (máximo peso total por envío)
        idx_dom = df_prod.groupby(COL_ENVIO)[COL_PESO].idxmax()
        df_dom = (
            df_prod.loc[idx_dom, [COL_ENVIO, COL_PROD, COL_PESO]]
            .rename(columns={COL_PROD: "PRODUCTO_DOMINANTE", COL_PESO: "PESO_DOMINANTE"})
        )
    
        # Lista (ordenada por peso desc, ya sin repetidos)
        df_list = (
            df_prod.sort_values([COL_ENVIO, COL_PESO], ascending=[True, False])
            .groupby(COL_ENVIO, as_index=False)[COL_PROD]
            .apply(lambda s: " | ".join(s.head(50).tolist()))
            .rename(columns={COL_PROD: "PRODUCTOS_LISTA"})
        )
    
        df_prod_agg = df_dom.merge(df_list, on=COL_ENVIO, how="left")
    
        # ----- cruzar solo los que NO tienen contenido -----
        df_concat_copy = df_concat.copy()
    
        df_concat_copy["CONTENIDO"] = df_concat_copy.get(
            "CONTENIDO", pd.Series([pd.NA] * len(df_concat_copy))
        ).astype("string")
    
        mask_sin_contenido = df_concat_copy["CONTENIDO"].isna() | (df_concat_copy["CONTENIDO"].str.strip() == "")
    
        df_to_ai = df_concat_copy.loc[mask_sin_contenido].copy()
        df_to_ai["guia"] = _clean_str_series(df_to_ai["guia"])
    
        df_to_ai = df_to_ai.merge(
            df_prod_agg,
            how="left",
            left_on="guia",
            right_on=COL_ENVIO
        )
    
        df_concat_copy.loc[mask_sin_contenido, "PRODUCTO_DOMINANTE"] = df_to_ai["PRODUCTO_DOMINANTE"].values
        df_concat_copy.loc[mask_sin_contenido, "PESO_DOMINANTE"] = df_to_ai["PESO_DOMINANTE"].values
        df_concat_copy.loc[mask_sin_contenido, "PRODUCTOS_LISTA"] = df_to_ai["PRODUCTOS_LISTA"].values
    
        def get_openai_client() -> OpenAI:
            return OpenAI(api_key=st.secrets["openai"]["api_key"])
    
    
        CATEGORIAS = [
            "Tenis", "Calzado", "Celular", "Computador", "Componente de computador",
            "Ropa", "Perfumes", "Cosmeticos", "Accesorios", "Reloj/Joyeria",
            "Hogar", "Electrodomestico", "Juguetes", "Herramientas",
            "Suplementos", "Medicamentos", "Alimentos",
            "Libros/Papeleria", "Miscelaneo"
        ]
    
        # ✅ ESTE es el JSON Schema REAL (raíz type=object)
        CLASIFICAR_ENVIO_SCHEMA = {
            "type": "object",
            "properties": {
                "categoria": {"type": "string", "enum": CATEGORIAS},
                "confianza": {"type": "integer", "minimum": 0, "maximum": 100},
                "contenido": {
                    "type": "string",
                    "description": "Texto corto para CONTENIDO (sin marcas). Ej: 'Calzado (tenis)', 'Celular', 'Ropa'."
                }
            },
            "required": ["categoria", "confianza", "contenido"],
            "additionalProperties": False
        }
    
        def gpt_clasificar_envio(client: OpenAI, producto_dominante: str, productos_lista: str) -> dict:
            prompt = f"""
        Eres un clasificador de envíos para manifiestos.
        Objetivo: llenar la columna CONTENIDO con una categoria GENERAL, sin marcas.
    
        Reglas:
        - Prioriza PRODUCTO_DOMINANTE (es el más pesado del envío).
        - Usa PRODUCTOS_LISTA solo como contexto.
        - NO menciones marcas (Nike, Apple, etc). Solo categoria general.
        - 'contenido' debe ser corto (1-4 palabras). Puedes aclarar entre paréntesis sin marcas.
        - Si es muy variado, usa Miscelaneo.
    
        PRODUCTO_DOMINANTE: {producto_dominante}
        PRODUCTOS_LISTA: {productos_lista}
        """.strip()
    
            resp = client.responses.create(
                model="gpt-4o-mini",
                input=prompt,
                text={
                    "format": {
                        "type": "json_schema",
                        "name": "clasificar_envio",   # <=64 chars
                        "schema": CLASIFICAR_ENVIO_SCHEMA,
                        "strict": True
                    }
                }
            )
    
            # ✅ resp.output_text ya es JSON válido
            return json.loads(resp.output_text)   

        client = get_openai_client()
    
        df_concat_copy["CONTENIDO"] = df_concat_copy.get(
            "CONTENIDO", pd.Series([pd.NA] * len(df_concat_copy))
        ).astype("string")
    
        mask_sin_contenido = df_concat_copy["CONTENIDO"].isna() | (df_concat_copy["CONTENIDO"].str.strip() == "")
    
        mask_listo_para_gpt = (
            mask_sin_contenido
            & df_concat_copy["PRODUCTO_DOMINANTE"].notna()
            & (df_concat_copy["PRODUCTO_DOMINANTE"].astype(str).str.strip() != "")
        )
    
        cache = {}
    
        cats, confs, conts = [], [], []
    
        for dom, lista in zip(
            df_concat_copy.loc[mask_listo_para_gpt, "PRODUCTO_DOMINANTE"].astype(str).tolist(),
            df_concat_copy.loc[mask_listo_para_gpt, "PRODUCTOS_LISTA"].fillna("").astype(str).tolist()
        ):
            lista = lista[:3000]  # límite de texto
            key = (dom, lista)
    
            if key in cache:
                out = cache[key]
            else:
                out = gpt_clasificar_envio(client, dom, lista)
                cache[key] = out
    
            cats.append(out["categoria"])
            confs.append(out["confianza"])
            conts.append(out["contenido"])
    
        df_concat_copy.loc[mask_listo_para_gpt, "CATEGORIA_GPT"] = cats
        df_concat_copy.loc[mask_listo_para_gpt, "CONF_GPT"] = confs
        df_concat_copy.loc[mask_listo_para_gpt, "CONTENIDO"] = conts
    
        # Si sigue sin match (no había productos), se queda vacío

        # =========================
        # PASAR CONTENIDO de df_concat_copy -> df_concat (por guia)
        # =========================
    
        # 1) Normalizar llaves (por si acaso)
        df_concat_copy["guia"] = _clean_str_series(df_concat_copy["guia"])
        df_concat["guia"] = _clean_str_series(df_concat["guia"])
    
        # 2) Asegurar CONTENIDO como string
        df_concat["CONTENIDO"] = df_concat.get("CONTENIDO", pd.Series([pd.NA]*len(df_concat))).astype("string")
        df_concat_copy["CONTENIDO"] = df_concat_copy.get("CONTENIDO", pd.Series([pd.NA]*len(df_concat_copy))).astype("string")
    
        # 3) Tomar SOLO filas donde df_concat_copy tiene contenido generado (no vacío)
        mask_copy_con = df_concat_copy["CONTENIDO"].notna() & (df_concat_copy["CONTENIDO"].str.strip() != "")
    
        # Si hay duplicados de guia en copy, me quedo con el último no vacío (por seguridad)
        df_map = (
            df_concat_copy.loc[mask_copy_con, ["guia", "CONTENIDO"]]
            .drop_duplicates(subset=["guia"], keep="last")
        )
    
        map_contenido = df_map.set_index("guia")["CONTENIDO"].to_dict()
    
        # 4) POBLAR en df_concat SOLO donde estaba vacío
        mask_concat_vacio = df_concat["CONTENIDO"].isna() | (df_concat["CONTENIDO"].str.strip() == "")
    
        df_concat.loc[mask_concat_vacio, "CONTENIDO"] = df_concat.loc[mask_concat_vacio, "guia"].map(map_contenido)
        
 
    

        resumen_manifiestos = [
            f"Nuevo 11591={nuevo_man_11591}",
            f"Otros={nuevo_man_otros}",
        ]
        if nuevo_man_contingencia is not None:
            resumen_manifiestos.insert(0, f"Contingencia={nuevo_man_contingencia}")

        st.success(f"Manifiestos asignados. {' | '.join(resumen_manifiestos)}")

        # -----------------------------
        # 12) Subir histórico actualizado (overwrite) a Dropbox
        # -----------------------------
        guias_contingencia_reales = {
            _normalize_excel_key(guia)
            for guia in df_concat.loc[df_concat["CONTINGENCIA_REAL"].fillna(False), "guia"].tolist()
            if _normalize_excel_key(guia)
        }

        excel_bytes = dfs_to_excel_bytes(
            {
                "HISTORICO": df_concat,
                "costo por manifiesto": df_costos_manifiesto,
            },
            highlight_rows={
                "HISTORICO": {"guia_col": "guia", "guias": guias_contingencia_reales},
            },
            hidden_columns={
                "HISTORICO": ["CONTINGENCIA_REAL"],
            },
        )
    
        dbx.files_upload(excel_bytes, DBX_FILE_PATH, mode=dropbox.files.WriteMode.overwrite)

        st.success("Histórico actualizado en Dropbox (reemplazado) ✅")

        # (opcional) mostrar muestra
        with st.expander("Ver muestra del histórico resultante"):
            st.dataframe(df_concat.drop(columns=["CONTINGENCIA_REAL"], errors="ignore").head(100))
        st.session_state["df_concat"] = df_concat
        st.session_state["fecha_str"] = datetime.now(ZoneInfo("America/New_York")).strftime("%Y-%m-%d")
        
        
        
        
    import zipfile

    st.divider()
    st.subheader("Descargas por manifiesto")

    if "df_concat" not in st.session_state:
        st.info("Primero ejecuta: **Procesar y actualizar histórico en Dropbox** para habilitar descargas.")
    else:
        df_concat = st.session_state["df_concat"]
        fecha_str = st.session_state.get("fecha_str") or datetime.now(ZoneInfo("America/New_York")).strftime("%Y-%m-%d")

        # -----------------------------
        # Vista SOLO para descargas (no afecta Dropbox)
        # - omite CASILLERO
        # - ordena columnas
        # - renombra guia -> GUIA
        # -----------------------------
        guias_contingencia_reales = {
            _normalize_excel_key(guia)
            for guia in df_concat.loc[df_concat["CONTINGENCIA_REAL"].fillna(False), "guia"].tolist()
            if _normalize_excel_key(guia)
        }

        df_dl = df_concat.copy()
        df_dl = df_dl.drop(columns=["CASILLERO","VALOR","CONTINGENCIA_REAL"], errors="ignore")

        if "guia" in df_dl.columns:
            df_dl = df_dl.rename(columns={"guia": "GUIA"})

        orden_descarga = [
            "FECHA GUIA",
            "GUIA",
            "COMPAÑÍA REMITENTE",
            "REMITENTE DIRECCION",
            "REMITENTE TELEFONO",
            "REMITENTE CIUDAD",
            "REMITENTE ESTADO",
            "NOMBRE DESTINO",
            "DESTINO DIRECCION",
            "DESTINO TELEFONO",
            "DESTINO CIUDAD",
            "CONTENIDO",
            "PESO LIBRAS",
            "PESO KILOS",
            "VALOR DECLARADO",
            "PIEZAS",
            "DESTINO ESTADO",
            "POSICION ARANCELARIA",
            "MANIFIESTO",
            "INSTRUCCIONES"
        ]

        presentes = [c for c in orden_descarga if c in df_dl.columns]
        extras = [c for c in df_dl.columns if c not in presentes]
        df_dl = df_dl[presentes + extras]

        # Lista de manifiestos disponibles (ordenados, sin nulos)
        if "MANIFIESTO" not in df_dl.columns:
            st.warning("No existe la columna MANIFIESTO en el histórico para descargas.")
        else:
            manifiestos = (
                df_dl["MANIFIESTO"]
                .dropna()
                .astype("int64")
                .sort_values()
                .unique()
                .tolist()
            )


            def resumen_costo_por_manifiesto(df: pd.DataFrame) -> pd.DataFrame:
                """
                Suma PESO LIBRAS, PESO KILOS, PIEZAS, COSTO por MANIFIESTO (solo para el df recibido).
                """
                df2 = df.copy()
        
                for col in ["PESO LIBRAS", "PESO KILOS", "PIEZAS", "COSTO"]:
                    if col in df2.columns:
                        df2[col] = pd.to_numeric(df2[col], errors="coerce")
        
                if "MANIFIESTO" not in df2.columns:
                    return pd.DataFrame(columns=["MANIFIESTO", "PESO LIBRAS", "PESO KILOS", "PIEZAS", "COSTO"])
        
                out = (
                    df2
                    .groupby("MANIFIESTO", as_index=False)[["PESO LIBRAS", "PESO KILOS", "PIEZAS", "COSTO"]]
                    .sum(min_count=1)
                )
                return out
        
        
            def build_zip_all_manifiestos(df_all: pd.DataFrame) -> bytes:
                zip_buf = io.BytesIO()
                with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
                    for man in (
                        df_all["MANIFIESTO"].dropna().astype("int64").sort_values().unique().tolist()
                    ):
                        df_m = df_all[df_all["MANIFIESTO"].astype("Int64") == man].copy()
        
                        df_costos_m = resumen_costo_por_manifiesto(df_m)
        
                        guias_rojas_man = {
                            _normalize_excel_key(guia)
                            for guia in df_m.get("GUIA", pd.Series(dtype="string")).tolist()
                            if _normalize_excel_key(guia) in guias_contingencia_reales
                        }

                        excel_bytes = dfs_to_excel_bytes(
                            {
                                f"MAN_{man}": df_m,
                                "costo por manifiesto": df_costos_m,
                            },
                            highlight_rows={
                                f"MAN_{man}": {"guia_col": "GUIA", "guias": guias_rojas_man},
                            },
                        )
        
                        filename = f"{fecha_str}-{man}.xlsx"
                        zf.writestr(filename, excel_bytes)
        
                zip_buf.seek(0)
                return zip_buf.getvalue()
        
        
            col_all, col_one = st.columns([1, 1])
        
            with col_all:
                st.markdown("**Descargar todos (ZIP)**")
                if st.button("Preparar ZIP con todos los manifiestos"):
                    zip_bytes = build_zip_all_manifiestos(df_dl)
                    st.download_button(
                        "Descargar ZIP",
                        data=zip_bytes,
                        file_name=f"{fecha_str}-manifiestos.zip",
                        mime="application/zip",
                    )
        
            with col_one:
                st.markdown("**Descargar uno puntual**")
                if not manifiestos:
                    st.info("No hay manifiestos disponibles para descargar.")
                else:
                    man_sel = st.selectbox("Buscar/seleccionar manifiesto", options=manifiestos)
        
                    df_sel = df_dl[df_dl["MANIFIESTO"].astype("Int64") == int(man_sel)].copy()
                    df_costos_sel = resumen_costo_por_manifiesto(df_sel)
        
                    guias_rojas_sel = {
                        _normalize_excel_key(guia)
                        for guia in df_sel.get("GUIA", pd.Series(dtype="string")).tolist()
                        if _normalize_excel_key(guia) in guias_contingencia_reales
                    }

                    excel_sel = dfs_to_excel_bytes(
                        {
                            f"MAN_{man_sel}": df_sel,
                            "costo por manifiesto": df_costos_sel,
                        },
                        highlight_rows={
                            f"MAN_{man_sel}": {"guia_col": "GUIA", "guias": guias_rojas_sel},
                        },
                    )
        
                    st.download_button(
                        f"Descargar {fecha_str}-{man_sel}.xlsx",
                        data=excel_sel,
                        file_name=f"{fecha_str}-{man_sel}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

elif modo == "Celulares Fénix":
    st.subheader("Celulares Fénix")

    fx_col1, fx_col2, fx_col3 = st.columns(3)
    with fx_col1:
        up_a_fx = st.file_uploader("A — Pistoleo Fénix (.xlsx)", type=["xlsx", "xls"], key="up_a_fx")
    with fx_col2:
        up_b_fx = st.file_uploader("B — Envíos Encargomio (.xlsx)", type=["xlsx", "xls"], key="up_b_fx")
    with fx_col3:
        up_p_fx = st.file_uploader("P — Productos por casillero (.xlsx)", type=["xlsx", "xls"], key="up_p_fx")

    run_fx = st.button("Procesar Celulares Fénix")

    if run_fx:
        # a) Validar que los 3 archivos estén cargados
        if not (up_a_fx and up_b_fx and up_p_fx):
            st.warning("Sube los 3 archivos (A Pistoleo Fénix, B Envíos Encargomio y P Productos) para continuar.")
            st.stop()

        # b) Leer los 3 archivos
        df_a_fx = pd.read_excel(up_a_fx)
        df_b_fx = pd.read_excel(up_b_fx)
        df_p_fx = pd.read_excel(up_p_fx)

        # c) Limpiar llaves con la función global _clean_str_series
        df_a_fx["Guias fenix"] = _clean_str_series(df_a_fx["Guias fenix"])
        df_b_fx["GUIA INTERNACIONAL"] = _clean_str_series(df_b_fx["GUIA INTERNACIONAL"])
        df_b_fx["NUMERO ENVIO"] = _clean_str_series(df_b_fx["NUMERO ENVIO"])
        df_p_fx["Envío"] = _clean_str_series(df_p_fx["Envío"])

        # d) Deduplicar A por guía (manda lo pistoleado)
        df_a_fx = df_a_fx.drop_duplicates(subset=["Guias fenix"])

        # e) CRUCE 1: A ⟕ B por guía (how="left": manda lo pistoleado)
        df_ab = df_a_fx.merge(
            df_b_fx,
            how="left",
            left_on="Guias fenix",
            right_on="GUIA INTERNACIONAL",
        )

        # f) Guías pistoleadas sin match en B (informativo, no detiene)
        mask_sin_match_fx = df_ab["NUMERO ENVIO"].isna()
        guias_sin_match_fx = (
            df_ab.loc[mask_sin_match_fx, "Guias fenix"].dropna().astype(str).tolist()
        )
        if guias_sin_match_fx:
            st.warning(
                "Guías pistoleadas sin match en Envíos: "
                + ", ".join(guias_sin_match_fx)
            )

        # g) CRUCE 2: con P por número de envío ("NUMERO ENVIO" de B ↔ "Envío" de P)
        # La columna "Envío" de Productos es numérica y contiene NUMERO ENVIO, no la guía de Fénix.
        df_ab["NUMERO ENVIO"] = _clean_str_series(df_ab["NUMERO ENVIO"])
        df_full = df_ab.merge(
            df_p_fx,
            how="left",
            left_on="NUMERO ENVIO",
            right_on="Envío",
        )

        # h) Guardar y verificación visual temporal
        st.session_state["df_fenix"] = df_full

        total_pistoleadas_fx = df_a_fx["Guias fenix"].nunique()
        total_con_b_fx = (~mask_sin_match_fx).sum()
        total_filas_producto_fx = len(df_full)

        st.success(
            f"Pistoleadas: {total_pistoleadas_fx} | "
            f"Cruzaron con B: {total_con_b_fx} | "
            f"Filas de producto resultantes: {total_filas_producto_fx}"
        )
        st.dataframe(df_full.head(50), use_container_width=True)

        # -----------------------------
        # 1) Renombrado de B (replica Luma, pero informativo: no st.stop)
        # -----------------------------
        rename_map = {
            "CLIENTE DESTINO": "NOMBRE DESTINO",
            "CIUDAD DESTINO": "DESTINO CIUDAD",
            "DEPARTAMENTO DESTINO": "DESTINO ESTADO",
            "PESO": "PESO LIBRAS",
        }
        alias_cols = {
            "DIRECCIÓN DESTINO": "DESTINO DIRECCION",
            "DIRECCIÃ“N DESTINO": "DESTINO DIRECCION",
            "TELÉFONO": "DESTINO TELEFONO",
            "TELÃ‰FONO": "DESTINO TELEFONO",
        }

        for src, dst in alias_cols.items():
            if src in df_full.columns:
                df_full = df_full.rename(columns={src: dst})

        faltantes_fx = [c for c in rename_map.keys() if c not in df_full.columns]
        if faltantes_fx:
            st.warning(f"Columnas de destinatario ausentes en B (se crean vacías): {faltantes_fx}")

        df_full = df_full.rename(columns=rename_map)

        # 2) Columna guia (la guía de Fénix es la llave de salida)
        df_full["guia"] = df_full["Guias fenix"]

        # 3) Columnas calculadas (replica Luma, sin POSICION ARANCELARIA ni COSTO)
        df_full["PIEZAS"] = 1
        df_full["VALOR DECLARADO"] = np.random.randint(91, 100, size=len(df_full))
        df_full["PESO LIBRAS"] = pd.to_numeric(df_full["PESO LIBRAS"], errors="coerce")
        df_full["PESO KILOS"] = df_full["PESO LIBRAS"] / 2.20462
        hoy_miami = datetime.now(ZoneInfo("America/New_York")).date()
        df_full["FECHA GUIA"] = pd.to_datetime(hoy_miami)

        # 4) CONTENIDO fijo (sin IA)
        df_full["CONTENIDO"] = "Celulares"

        # 5) Agregación a una fila por guía (productos concatenados)
        productos_por_guia = (
            df_full.dropna(subset=["Nombre producto"])
            .groupby("guia")["Nombre producto"]
            .apply(lambda s: " | ".join(s.astype(str).unique()))
            .rename("PRODUCTOS")
        )

        df_guia = df_full.drop_duplicates(subset=["guia"], keep="first").copy()
        df_guia = df_guia.merge(productos_por_guia, how="left", on="guia")

        # 6) Esquema final de salida (crea vacías las que falten)
        cols_salida_fx = [
            "FECHA GUIA", "guia", "NOMBRE DESTINO", "DESTINO DIRECCION",
            "DESTINO TELEFONO", "DESTINO CIUDAD", "DESTINO ESTADO", "CONTENIDO",
            "PRODUCTOS", "PESO LIBRAS", "PESO KILOS", "VALOR DECLARADO", "PIEZAS",
            "FECHA_PROCESO", "LOTE",
        ]
        for c in cols_salida_fx:
            if c not in df_guia.columns:
                df_guia[c] = pd.NA
        df_salida_fx = df_guia[cols_salida_fx].copy()

        # 6.1) Cargar histórico Fénix, calcular LOTE y sellar las guías nuevas
        #      (se hace ANTES de mostrar el detalle para que FECHA_PROCESO y LOTE
        #       aparezcan en pantalla y en el Excel de descarga completa)
        DBX_FILE_PATH_FENIX = "/Manifiestos/Celulares_fenix.xlsx"

        def load_historico_fenix(dbx, path):
            try:
                _, res = dbx.files_download(path)
                return pd.read_excel(io.BytesIO(res.content), sheet_name=0)
            except dropbox.exceptions.ApiError:
                return pd.DataFrame(columns=df_salida_fx.columns)

        dbx_fx = get_dbx()
        hist_fx = load_historico_fenix(dbx_fx, DBX_FILE_PATH_FENIX)

        # Normalizar llave "guia" en ambos (por .0 / espacios) con la global
        if "guia" in hist_fx.columns and not hist_fx.empty:
            hist_fx["guia"] = _clean_str_series(hist_fx["guia"])
        df_salida_fx["guia"] = _clean_str_series(df_salida_fx["guia"])

        # Calcular LOTE de esta corrida (después de cargar hist_fx, antes de concatenar)
        if (not hist_fx.empty) and ("LOTE" in hist_fx.columns):
            _max_lote_fx = pd.to_numeric(hist_fx["LOTE"], errors="coerce").max()
            lote_actual = int(_max_lote_fx) + 1 if pd.notna(_max_lote_fx) else 1
        else:
            lote_actual = 1

        # Sellar SOLO las guías nuevas (df_salida_fx); las del histórico ya traen su sello.
        # OJO: el datetime debe quedar tz-naive (Excel/openpyxl no soporta tz),
        # conservando la hora local de Miami.
        now_miami = datetime.now(ZoneInfo("America/New_York"))
        fecha_proceso_naive = now_miami.replace(microsecond=0, tzinfo=None)
        df_salida_fx["FECHA_PROCESO"] = pd.to_datetime(fecha_proceso_naive)
        df_salida_fx["LOTE"] = lote_actual

        # Salvaguarda: si algún histórico viejo trae FECHA_PROCESO tz-aware,
        # la pasamos a tz-naive justo antes de escribir Excel (openpyxl no soporta tz).
        def _strip_tz(df):
            if "FECHA_PROCESO" in df.columns:
                s = pd.to_datetime(df["FECHA_PROCESO"], errors="coerce")
                if getattr(s.dt, "tz", None) is not None:
                    df = df.copy()
                    df["FECHA_PROCESO"] = s.dt.tz_localize(None)
            return df

        # 7) Resumen agregado
        total_guias_fx = df_salida_fx["guia"].nunique()
        total_piezas_fx = pd.to_numeric(df_salida_fx["PIEZAS"], errors="coerce").sum()
        total_libras_fx = pd.to_numeric(df_salida_fx["PESO LIBRAS"], errors="coerce").sum()
        total_kilos_fx = pd.to_numeric(df_salida_fx["PESO KILOS"], errors="coerce").sum()

        guias_por_ciudad_fx = (
            df_salida_fx.groupby("DESTINO CIUDAD", dropna=False)["guia"]
            .nunique()
            .reset_index(name="GUIAS")
            .sort_values("GUIAS", ascending=False)
        )

        # 8) Pantalla: detalle
        st.subheader("Detalle Celulares")
        st.dataframe(df_salida_fx, use_container_width=True)

        st.markdown("### Resumen Celulares Fénix")
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Guías (envíos)", int(total_guias_fx))
        m2.metric("Piezas", int(total_piezas_fx))
        m3.metric("Peso libras", round(float(total_libras_fx), 2))
        m4.metric("Peso kilos", round(float(total_kilos_fx), 2))

        st.markdown("**Guías por ciudad**")
        st.dataframe(guias_por_ciudad_fx, use_container_width=True)

        # 9) Descarga Excel (reusa la función global dfs_to_excel_bytes)
        excel_fx = dfs_to_excel_bytes({
            "CELULARES": _strip_tz(df_salida_fx),
            "RESUMEN": _strip_tz(guias_por_ciudad_fx),
        })
        st.download_button(
            label="Descargar Excel Celulares Fénix",
            data=excel_fx,
            file_name=f"celulares_fenix_{datetime.now(ZoneInfo('America/New_York')).date()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # 10) Guardar para el siguiente paso (Dropbox)
        st.session_state["df_salida_fx"] = df_salida_fx

        # -----------------------------
        # 11) Persistir histórico Fénix en Dropbox (archivo SEPARADO del de Luma)
        # -----------------------------
        st.caption(
            "El histórico de Celulares Fénix se guarda en un archivo aparte del de Luma. "
            "Esta acción SOBREESCRIBE ese archivo de histórico Fénix en Dropbox."
        )

        # (El histórico Fénix ya se cargó en el paso 6.1; DBX_FILE_PATH_FENIX, hist_fx
        #  y lote_actual están definidos. df_salida_fx ya viene sellado.)

        # Si el histórico viene de antes de este cambio (sin FECHA_PROCESO / LOTE),
        # crear esas columnas para que el concat no desalinee
        if "FECHA_PROCESO" not in hist_fx.columns:
            hist_fx["FECHA_PROCESO"] = pd.NaT
        if "LOTE" not in hist_fx.columns:
            hist_fx["LOTE"] = pd.NA

        # Concatenar y deduplicar conservando lo VIEJO (histórico manda)
        df_hist_fx = pd.concat([hist_fx, df_salida_fx], ignore_index=True)
        df_hist_fx = df_hist_fx.drop_duplicates(subset=["guia"], keep="first")

        # Generar Excel del histórico con la global
        excel_hist_fx = dfs_to_excel_bytes({"HISTORICO": _strip_tz(df_hist_fx)})

        # Subir a Dropbox sobreescribiendo (mismo patrón que Luma), con manejo de error
        try:
            dbx_fx.files_upload(
                excel_hist_fx,
                DBX_FILE_PATH_FENIX,
                mode=dropbox.files.WriteMode.overwrite,
            )
            st.success(
                f"Histórico Fénix actualizado en Dropbox: {len(df_hist_fx)} guías acumuladas "
                f"({DBX_FILE_PATH_FENIX})."
            )
        except Exception as e:
            st.error(f"No se pudo actualizar el histórico Fénix en Dropbox: {e}")

        # 12) Descargar SOLO el lote de esta corrida (lo realmente nuevo).
        #     Las guías que ya existían conservaron su lote viejo (keep="first"),
        #     por lo que no aparecen aquí: es exactamente "solo lo nuevo".
        df_lote_actual = df_hist_fx[df_hist_fx["LOTE"] == lote_actual].copy()
        excel_lote = dfs_to_excel_bytes({"LOTE_ACTUAL": _strip_tz(df_lote_actual)})
        st.download_button(
            label=f"Descargar solo lo nuevo (lote {lote_actual})",
            data=excel_lote,
            file_name=f"celulares_fenix_lote{lote_actual}_{now_miami.date()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.info(
            f"Esta corrida agregó {len(df_lote_actual)} guías nuevas (lote {lote_actual})."
        )

elif modo == "Envios Encargomio":
    st.subheader("Envios Encargomio — formato cajas (histórico propio)")
    st.caption(
        "Cruza Pistoleo × Envíos Encargomio, clasifica CONTENIDO con IA (la key de "
        "OpenAI se toma en línea), numera manifiestos desde 2000001 (+1 por corrida) "
        "y guarda en un histórico APARTE en Dropbox. La descarga arma los totales "
        "(W.Lb / W.Kg / Value / Pieces) de forma automática por cada manifiesto."
    )

    DBX_FILE_PATH_EE = "/Manifiestos/Envios_encargomio.xlsx"

    ee1, ee2, ee3 = st.columns(3)
    with ee1:
        up_a_ee = st.file_uploader("A — Pistoleo (Envio)", type=["xlsx", "xls"], key="up_a_ee")
    with ee2:
        up_b_ee = st.file_uploader("B — Envíos Encargomio (NUMERO ENVIO)", type=["xlsx", "xls"], key="up_b_ee")
    with ee3:
        up_p_ee = st.file_uploader("P — Productos (para CONTENIDO por IA)", type=["xlsx", "xls"], key="up_p_ee")

    run_ee = st.button(
        "Procesar y actualizar histórico Envios Encargomio",
        type="primary",
        disabled=not (up_a_ee and up_b_ee),
    )

    if run_ee:
        dbx = get_dbx()
        try:
            _, res = dbx.files_download(DBX_FILE_PATH_EE)
            hist_ee = pd.read_excel(io.BytesIO(res.content), sheet_name=0)
        except dropbox.exceptions.ApiError:
            hist_ee = pd.DataFrame()

        df_a_ee = pd.read_excel(up_a_ee)
        df_b_ee = pd.read_excel(up_b_ee)
        df_p_ee = pd.read_excel(up_p_ee) if up_p_ee is not None else None

        df_concat_ee, nuevo_man, n_ia, guias_sin_peso, guias_sin_categoria = procesar_envios_encargomio(
            df_a_ee, df_b_ee, hist_ee, df_p_ee
        )

        excel_hist_ee = dfs_to_excel_bytes({"HISTORICO": df_concat_ee})
        dbx.files_upload(
            excel_hist_ee, DBX_FILE_PATH_EE, mode=dropbox.files.WriteMode.overwrite
        )

        st.session_state["hist_ee"] = df_concat_ee
        if n_ia >= 0:
            msg_ia = f"CONTENIDO por IA: {n_ia} guía(s)"
        else:
            msg_ia = "IA no disponible en esta corrida (se conservó el CONTENIDO existente)"
        st.success(
            f"Histórico Envios Encargomio actualizado ({len(df_concat_ee)} filas). "
            f"Nuevo manifiesto: {nuevo_man}. {msg_ia}."
        )

        if guias_sin_peso:
            st.warning(
                f"⚠ {len(guias_sin_peso)} guía(s) pistoleadas SIN cruce en Envíos Encargomio "
                f"o con peso inválido. Entran al histórico con peso vacío; revísalas: "
                + ", ".join(guias_sin_peso)
            )

        if guias_sin_categoria:
            st.warning(
                f"⚠ {len(guias_sin_categoria)} envío(s) SIN categoría (venían 'Otro'/vacío y no "
                f"tienen producto en el casillero para reclasificar). Requieren CONTENIDO manual "
                f"para la DIAN: " + ", ".join(guias_sin_categoria)
            )

    if "hist_ee" not in st.session_state:
        if st.button("Cargar histórico Envios Encargomio (solo descargar)"):
            dbx = get_dbx()
            try:
                _, res = dbx.files_download(DBX_FILE_PATH_EE)
                st.session_state["hist_ee"] = pd.read_excel(io.BytesIO(res.content), sheet_name=0)
                st.success("Histórico cargado.")
            except dropbox.exceptions.ApiError:
                st.warning("Aún no existe el histórico Envios Encargomio. Procesa una vez para crearlo.")

    if "hist_ee" in st.session_state:
        hist_ee = st.session_state["hist_ee"]
        if "guia" not in hist_ee.columns and "GUIA" in hist_ee.columns:
            hist_ee = hist_ee.rename(columns={"GUIA": "guia"})

        st.divider()
        st.subheader("Descargar manifiesto (formato cajas con totales)")

        if "MANIFIESTO" not in hist_ee.columns:
            st.warning("El histórico no tiene columna MANIFIESTO.")
        else:
            manifiestos = (
                pd.to_numeric(hist_ee["MANIFIESTO"], errors="coerce")
                .dropna()
                .astype("int64")
                .sort_values()
                .unique()
                .tolist()
            )
            if not manifiestos:
                st.info("El histórico no tiene manifiestos disponibles.")
            else:
                man_sel = st.selectbox("Selecciona el manifiesto", options=manifiestos)

                df_m = hist_ee[
                    pd.to_numeric(hist_ee["MANIFIESTO"], errors="coerce") == int(man_sel)
                ].copy()
                df_info = build_info_manifiesto_df(df_m)

                tot_lb = round(float(pd.to_numeric(df_info["PESO LIBRAS"], errors="coerce").sum()), 2)
                tot_kg = round(float(pd.to_numeric(df_info["PESO KILOS"], errors="coerce").sum()), 2)
                tot_pz = int(pd.to_numeric(df_info["PIEZAS"], errors="coerce").sum())
                tot_val = round(float(pd.to_numeric(df_info["VALOR DECLARADO"], errors="coerce").sum()), 2)

                st.markdown(f"### Manifiesto {int(man_sel)} — {df_info['GUIA / consignment'].nunique()} guías")
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Pieces", tot_pz)
                c2.metric("W.Lb", tot_lb)
                c3.metric("W.Kg", tot_kg)
                c4.metric("Value $", tot_val)

                st.dataframe(df_info, use_container_width=True)

                data_info = info_manifiesto_excel_bytes(df_info)
                fecha_str = datetime.now(ZoneInfo("America/New_York")).strftime("%Y-%m-%d")
                st.download_button(
                    label=f"Descargar {fecha_str}-{int(man_sel)}-CAJAS.xlsx",
                    data=data_info,
                    file_name=f"{fecha_str}-{int(man_sel)}-CAJAS.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
